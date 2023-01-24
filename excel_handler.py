from datetime import datetime
import openpyxl
from parsers.aquamobil_aqua import parser as aquamobil_aqua_def
from parsers.aquamobil_arhiz import parser as aquamobil_arhiz_def
from parsers.aquamobil_artenza import parser as aquamobil_artenza_def
from parsers.aquamobil_kukuzar import parser as aquamobil_kukuzar_def
from parsers.aquamobil_sosnovskaya import parser as aquamobil_sosnovskaya_def
from parsers.chebarkul import parser as chebarkul_def
from parsers.crystal import parser as crystal_def
from parsers.gorny import parser as gorny_def
from parsers.lubima import parser as lubima_def
from parsers.luxe_luxik import parser as luxe_luxik_def
from parsers.luxe import parser as luxe_def
from parsers.niagara_caucasus import parser as niagara_caucasus_def
from parsers.niagara_premium import parser as niagara_premium_def
from parsers.niagara import parser as niagara_def
from parsers.vlasov import parser as vlasov_def
from parsers.zhivaya import parser as zhivaya_def

async def excel_handle():
    wb = openpyxl.load_workbook('analysis.xlsx')
    month = datetime.now().month
    month = month if len(f'{month}') >= 2 else f'0{month}'
    year = f'{datetime.now().year}'[-2:]
    date = f'{datetime.now().day}.{month}.{year}'
    if date in wb:
        return
    last_sheet = wb.worksheets[-1]
    previous_date = last_sheet.title
    today_sheet = wb.copy_worksheet(last_sheet)
    today_sheet.title = date

    try:
        aquamobil_aqua = aquamobil_aqua_def()
    except:
        aquamobil_aqua = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        aquamobil_arhiz = aquamobil_arhiz_def()
    except:
        aquamobil_arhiz = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        aquamobil_artenza = aquamobil_artenza_def()
    except:
        aquamobil_artenza = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        aquamobil_kukuzar = aquamobil_kukuzar_def()
    except:
        aquamobil_kukuzar = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        aquamobil_sosnovskaya = aquamobil_sosnovskaya_def()
    except:
        aquamobil_sosnovskaya = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        chebarkul = chebarkul_def()
    except:
        chebarkul = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        crystal = crystal_def()
    except:
        crystal = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']
        
    try:
        gorny = gorny_def()
    except:
        gorny = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        lubima = lubima_def()
    except:
        lubima = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        luxe_luxik = luxe_luxik_def()
    except:
        luxe_luxik = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        luxe = luxe_def()
    except:
        luxe = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        niagara_caucasus = niagara_caucasus_def()
    except:
        niagara_caucasus = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        niagara_premium = niagara_premium_def()
    except:
        niagara_premium = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        niagara = niagara_def()
    except:
        niagara = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        vlasov = vlasov_def()
    except:
        vlasov = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        zhivaya = zhivaya_def()
    except:
        zhivaya = ['ошибка парсера', 'ошибка парсера', 'ошибка парсера']

    try:
        today_sheet.cell(row=2, column=4).value = aquamobil_aqua[0]
        today_sheet.cell(row=3, column=4).value = aquamobil_aqua[1]
    except:
        today_sheet.cell(row=2, column=4).value = 0
        today_sheet.cell(row=3, column=4).value = 0

    try:
        today_sheet.cell(row=2, column=5).value = aquamobil_arhiz[0]
        today_sheet.cell(row=3, column=5).value = aquamobil_arhiz[1]
    except:
        today_sheet.cell(row=2, column=5).value = 0
        today_sheet.cell(row=3, column=5).value = 0

    try:
        today_sheet.cell(row=2, column=6).value = aquamobil_artenza[0]
        today_sheet.cell(row=3, column=6).value = aquamobil_artenza[1]
    except:
        today_sheet.cell(row=2, column=6).value = 0
        today_sheet.cell(row=3, column=6).value = 0
    
    try:
        today_sheet.cell(row=2, column=7).value = aquamobil_kukuzar[0]
        today_sheet.cell(row=3, column=7).value = aquamobil_kukuzar[1]
    except:
        today_sheet.cell(row=2, column=7).value = 0
        today_sheet.cell(row=3, column=7).value = 0

    try:
        today_sheet.cell(row=2, column=8).value = aquamobil_sosnovskaya[0]
        today_sheet.cell(row=3, column=8).value = aquamobil_sosnovskaya[1]
    except:
        today_sheet.cell(row=2, column=8).value = 0
        today_sheet.cell(row=3, column=8).value = 0

    try:
        today_sheet.cell(row=2, column=9).value = chebarkul[0]
        today_sheet.cell(row=3, column=9).value = chebarkul[1]
        today_sheet.cell(row=8, column=9).value = chebarkul[2]
    except:
        today_sheet.cell(row=2, column=9).value = 0
        today_sheet.cell(row=3, column=9).value = 0
        today_sheet.cell(row=8, column=9).value = 0

    try:
        today_sheet.cell(row=3, column=10).value = crystal[1]
        today_sheet.cell(row=8, column=10).value = crystal[2]
    except:
        today_sheet.cell(row=3, column=10).value = 0
        today_sheet.cell(row=8, column=10).value = 0

    try:
        today_sheet.cell(row=3, column=11).value = gorny[1]
    except:
        today_sheet.cell(row=3, column=11).value = 0

    try:
        today_sheet.cell(row=2, column=12).value = lubima[0]
        today_sheet.cell(row=3, column=12).value = lubima[1]
        today_sheet.cell(row=8, column=12).value = lubima[2]
    except:
        today_sheet.cell(row=2, column=12).value = 0
        today_sheet.cell(row=3, column=12).value = 0
        today_sheet.cell(row=8, column=12).value = 0

    try:
        today_sheet.cell(row=2, column=13).value = luxe[0]
        today_sheet.cell(row=3, column=13).value = luxe[1]
        today_sheet.cell(row=8, column=13).value = luxe[2]
    except:
        today_sheet.cell(row=2, column=13).value = 0
        today_sheet.cell(row=3, column=13).value = 0
        today_sheet.cell(row=8, column=13).value = 0       

    try:
        today_sheet.cell(row=2, column=14).value = luxe_luxik[0]
        today_sheet.cell(row=3, column=14).value = luxe_luxik[1]
        today_sheet.cell(row=8, column=14).value = luxe_luxik[2]
    except:
        today_sheet.cell(row=2, column=14).value = 0
        today_sheet.cell(row=3, column=14).value = 0
        today_sheet.cell(row=8, column=14).value = 0    

    try:
        today_sheet.cell(row=2, column=15).value = niagara[0]
        today_sheet.cell(row=3, column=15).value = niagara[1]
        today_sheet.cell(row=8, column=15).value = niagara[2]
    except:
        today_sheet.cell(row=2, column=15).value = 0
        today_sheet.cell(row=3, column=15).value = 0
        today_sheet.cell(row=8, column=15).value = 0        

    try:
        today_sheet.cell(row=2, column=16).value = niagara_premium[0]
        today_sheet.cell(row=3, column=16).value = niagara_premium[1]
        today_sheet.cell(row=8, column=16).value = niagara_premium[2]
    except:
        today_sheet.cell(row=2, column=16).value = 0
        today_sheet.cell(row=3, column=16).value = 0
        today_sheet.cell(row=8, column=16).value = 0

    try:
        today_sheet.cell(row=2, column=17).value = niagara_caucasus[0]
        today_sheet.cell(row=3, column=17).value = niagara_caucasus[1]
        today_sheet.cell(row=8, column=17).value = niagara_caucasus[2]
    except:
        today_sheet.cell(row=2, column=17).value = 0
        today_sheet.cell(row=3, column=17).value = 0
        today_sheet.cell(row=8, column=17).value = 0

    try:
        today_sheet.cell(row=2, column=18).value = vlasov[0]
        today_sheet.cell(row=3, column=18).value = vlasov[1]
    except:
        today_sheet.cell(row=2, column=18).value = 0
        today_sheet.cell(row=3, column=18).value = 0

    try:
        today_sheet.cell(row=2, column=19).value = zhivaya[0]
        today_sheet.cell(row=3, column=19).value = zhivaya[1]
        today_sheet.cell(row=8, column=19).value = zhivaya[2]
    except:
        today_sheet.cell(row=2, column=19).value = 0
        today_sheet.cell(row=3, column=19).value = 0
        today_sheet.cell(row=8, column=19).value = 0

    today_sheet.cell(row=6, column=2).value = f"=B3-'{previous_date}'!B3"
    today_sheet.cell(row=6, column=4).value = f"=D3-'{previous_date}'!D3"
    today_sheet.cell(row=6, column=5).value = f"=E3-'{previous_date}'!E3"
    today_sheet.cell(row=6, column=6).value = f"=F3-'{previous_date}'!F3"
    today_sheet.cell(row=6, column=7).value = f"=G3-'{previous_date}'!G3"
    today_sheet.cell(row=6, column=8).value = f"=H3-'{previous_date}'!H3"
    today_sheet.cell(row=6, column=9).value = f"=I3-'{previous_date}'!I3"
    today_sheet.cell(row=6, column=10).value = f"=J3-'{previous_date}'!J3"
    today_sheet.cell(row=6, column=11).value = f"=K3-'{previous_date}'!K3"
    today_sheet.cell(row=6, column=12).value = f"=L3-'{previous_date}'!L3"
    today_sheet.cell(row=6, column=13).value = f"=M3-'{previous_date}'!M3"
    today_sheet.cell(row=6, column=14).value = f"=N3-'{previous_date}'!N3"
    today_sheet.cell(row=6, column=15).value = f"=O3-'{previous_date}'!O3"
    today_sheet.cell(row=6, column=16).value = f"=P3-'{previous_date}'!P3"
    today_sheet.cell(row=6, column=17).value = f"=Q3-'{previous_date}'!Q3"
    today_sheet.cell(row=6, column=18).value = f"=R3-'{previous_date}'!R3"
    today_sheet.cell(row=6, column=19).value = f"=S3-'{previous_date}'!S3"

    wb.save('analysis.xlsx')