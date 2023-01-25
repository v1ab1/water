import openpyxl
from datetime import datetime

async def graph_handle():
    wb = openpyxl.load_workbook('analysis.xlsx')
    month = datetime.now().month
    month = month if len(f'{month}') >= 2 else f'0{month}'
    year = f'{datetime.now().year}'[-2:]
    date = f'{datetime.now().day}.{month}.{year}'
    if date in wb:
        return
    my_file = open("index.txt", "r")
    index = int(my_file.read())
    my_file.close()
    index = index + 1
    my_file = open("update.txt", "w")
    my_file.write(f'{index}')
    my_file.close()
    sheet = wb.worksheets[0]
    sheet.cell(row=index, column=1).value = date
    sheet.cell(row=index, column=2).value = f"='{date}'!D3"
    sheet.cell(row=index, column=3).value = f"='{date}'!E3"
    sheet.cell(row=index, column=4).value = f"='{date}'!F3"
    sheet.cell(row=index, column=5).value = f"='{date}'!G3"
    sheet.cell(row=index, column=6).value = f"='{date}'!H3"
    sheet.cell(row=index, column=7).value = f"='{date}'!I3"
    sheet.cell(row=index, column=8).value = f"='{date}'!J3"
    sheet.cell(row=index, column=9).value = f"='{date}'!K3"
    sheet.cell(row=index, column=10).value = f"='{date}'!L3"
    sheet.cell(row=index, column=11).value = f"='{date}'!M3"
    sheet.cell(row=index, column=12).value = f"='{date}'!N3"
    sheet.cell(row=index, column=13).value = f"='{date}'!O3"
    sheet.cell(row=index, column=14).value = f"='{date}'!P3"
    sheet.cell(row=index, column=15).value = f"='{date}'!Q3"
    sheet.cell(row=index, column=16).value = f"='{date}'!R3"
    sheet.cell(row=index, column=17).value = f"='{date}'!S3"
    wb.save('analysis.xlsx')