import openpyxl

async def prices_handle():
    wb = openpyxl.load_workbook('./data/analysis.xlsx')
    last_sheet = wb.worksheets[-1]
    prices = [
        last_sheet.cell(row=3, column=4).value,
        last_sheet.cell(row=3, column=5).value,
        last_sheet.cell(row=3, column=6).value,
        last_sheet.cell(row=3, column=7).value,
        last_sheet.cell(row=3, column=8).value,
        last_sheet.cell(row=3, column=9).value,
        last_sheet.cell(row=3, column=10).value,
        last_sheet.cell(row=3, column=11).value,
        last_sheet.cell(row=3, column=12).value,
        last_sheet.cell(row=3, column=13).value,
        last_sheet.cell(row=3, column=14).value,
        last_sheet.cell(row=3, column=15).value,
        last_sheet.cell(row=3, column=16).value,
        last_sheet.cell(row=3, column=17).value,
        last_sheet.cell(row=3, column=18).value,
        last_sheet.cell(row=3, column=19).value
    ]
    return prices